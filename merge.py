from os import listdir
from os.path import isfile, join
from glob import glob
import  openpyxl
from openpyxl import Workbook

def get_all_value_excel(file_path):
    wb_obj = openpyxl.load_workbook(file_path)
    sheet_obj = wb_obj.active
    values = []
    for i in range(1,sheet_obj.max_row+1): # Ignore first row 0 value
        cell_obj = sheet_obj.cell(row=i, column=1)
        value=cell_obj.value
        if type(value)==type(0) and value==0: # ignore first cell if its 0
            continue
        values.append(value)
    wb_obj.close()
    return values

def get_multiple_excel_values(file_paths):
    values=[]
    for path in file_paths:
        values+=get_all_value_excel(path)
    return values

folder_names=[]
selector='Sectors'
excel_file_map={}
i=1
while True:
    folder_name=input(f"Pleasse enter {i} folder relative path : ")
    if folder_name.strip()!= '':
        folder_names.append(folder_name+'\\Sectors')
    else:
        continue
    i+=1
    more_folders=input("Do you have more folders (Y/N) ? ")
    if more_folders.strip().lower()=='y':
        continue
    else:
        break

for folder in folder_names:
    sectors=glob(folder+'/*/') # get sub folders of Sector. * means to get matching folders
    sectors=[s.replace('\\','/') for s in sectors]
    for sector in sectors:
        sector_name=sector.split('/')[-2]
        sector_folders=glob(sector+'*/') # get day folders in side sector
        sector_folders = [s.replace('\\', '/') for s in sector_folders]
        for day_folder in sector_folders:
            print(day_folder.replace('\\','/'))
            print(day_folder)

            days=int(day_folder.split('/')[-2].lower().replace('days','')) # get days
            onlyfiles = [join(day_folder, f) for f in listdir(day_folder) if isfile(join(day_folder, f))] # files in days folder
            print(onlyfiles)
            for file in onlyfiles:
                file_name=file.split('/')[-1]
                if ord(file_name[0])<65 or ord(file_name[0])>122: # if excel open a file it creates temp file. to skip that
                    continue
                if '.xlsx' in file_name and '_test.xlsx' not in file_name: # remove non excel files and excel files with test postfix
                    stock=file_name.split('_')[0].upper()
                    if sector_name in excel_file_map.keys():
                        if days in excel_file_map[sector_name].keys():
                            if stock in excel_file_map[sector_name][days].keys():
                                excel_file_map[sector_name][days][stock].append(file)
                            else:
                                excel_file_map[sector_name][days][stock]=[file]
                        else:
                            excel_file_map[sector_name][days]={stock:[file]}
                    else:
                        excel_file_map[sector_name]={days:{stock:[file]}}

print(f"File map : {excel_file_map}")
wb = Workbook()
sheets = list(excel_file_map.keys())
sheets.sort()
wb.active.title=sheets[0]

i=0
for sheet in sheets:
    days=list(excel_file_map[sheet].keys())
    days.sort()
    if i==0:
        ws=wb.active
    else:
        ws=wb.create_sheet(sheet)
    i+=1
    stock_colum_map={}
    last_column=65
    row=2 # keep track of start of day should be row
    for day in days:
        stocks=list(excel_file_map[sheet][day])
        stocks.sort()
        for stock in stocks: # make header columns
            if stock not in stock_colum_map.keys():
                ws[f'{chr(last_column)}1']=stock
                stock_colum_map[stock]=chr(last_column)
                last_column+=1
        max_values=0
        for stock in stocks:
            temp_row = row
            print(f'Processing Files {excel_file_map[sheet][day][stock]}')
            values=get_multiple_excel_values(excel_file_map[sheet][day][stock])
            for value in values:
                ws[f'{stock_colum_map[stock]}{temp_row}']=value
                temp_row+=1
            if max_values<len(values)+1+row:
                max_values=len(values)+1+row
        row=max_values

wb.save('test.xlsx')














